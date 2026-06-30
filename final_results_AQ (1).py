"""
final_results_AQ.py
=====================
Aggregates one or more detailed analysis Excel files (produced by
analysis_pipeline_AQ.run_analysis) into:
  - a simplified, wide-format summary (one row per sample, one column per
    assay, plus an overall call), with color-coded cells
  - a full-information long-format export with the key per-assay numbers

Sample-level controls (PC, NC, NTC) are excluded from both outputs.
"""

import os
import glob
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DESIRED_ASSAY_ORDER = [
    "TERT-124",
    "TERT-146",
    "FGFR3-248",
    "FGFR3-249",
    "FGFR3-372",
    "FGFR3-375",
]

CONTROL_SAMPLE_NAMES = {"PC", "NC", "NTC"}


def _is_positive(val):
    if isinstance(val, str):
        v = val.strip().lower()
        return v.startswith("positive")
    return False


def _is_inconclusive(val):
    if isinstance(val, str):
        return val.strip().lower() == "inconclusive"
    return False


def _is_negative(val):
    if isinstance(val, str):
        v = val.strip().lower()
        return v.startswith("negative")
    return False


class Resultados:

    def results_together(self, file_path_pattern):
        """Read and concatenate all detailed-analysis Excel files matching
        the given glob pattern into a single long-format DataFrame."""
        files = glob.glob(file_path_pattern)
        if not files:
            return None
        dfs = [pd.read_excel(f) for f in files]
        df_final = pd.concat(dfs, ignore_index=True)
        return df_final

    def select_information(self, df):
        """Keep only the columns needed for the simplified summary, and
        drop control samples (PC/NC/NTC)."""
        selected_columns = ["Source_File", "Sample description 1", "Assay", "Result"]
        df_selected = df[selected_columns].copy()
        df_selected = df_selected[
            ~df_selected["Sample description 1"].astype(str).str.strip().str.upper().isin(CONTROL_SAMPLE_NAMES)
        ]
        return df_selected

    def add_overall_result(self, df):
        """
        Per sample, across all its assays:
          - "Positive" if any assay is Positive (incl. '<20k events' variant)
          - "Inconclusive" if no Positive and any assay is Inconclusive
          - "Negative" otherwise
        Also lists which assays came back positive.
        """
        grouped = df.groupby("Sample description 1")
        overall = []
        for sample, group in grouped:
            positives = group[group["Result"].apply(_is_positive)]["Assay"].tolist()
            inconclusives = group[group["Result"].apply(_is_inconclusive)]["Assay"].tolist()

            if positives:
                overall_result = "Positive"
            elif inconclusives:
                overall_result = "Inconclusive"
            else:
                overall_result = "Negative"

            overall.append({
                "Sample description 1": sample,
                "Overall Result": overall_result,
                "Positive Assays": ", ".join(positives) if positives else "",
            })
        return pd.DataFrame(overall)

    def save_selected_information(self, df, filename, desired_order=None):
        """
        Save the simplified, wide-format summary: one row per sample, one
        column per assay (Result), plus Overall Result / Positive Assays.
        Color-codes each assay cell: green=Positive, red=Negative,
        yellow=Inconclusive.
        """
        required = ["Sample description 1", "Assay", "Result"]
        for col in required:
            if col not in df.columns:
                raise ValueError(f"DataFrame is missing required column: {col}")

        if desired_order is None:
            desired_order = DESIRED_ASSAY_ORDER

        if df.duplicated(subset=["Sample description 1", "Assay"]).any():
            df = df.groupby(["Sample description 1", "Assay"], as_index=False).first()

        wide_df = df.pivot(index="Sample description 1", columns="Assay", values="Result").reset_index()

        summary_df = self.add_overall_result(df)
        wide_df = wide_df.merge(summary_df, on="Sample description 1", how="left")

        for assay in desired_order:
            if assay not in wide_df.columns:
                wide_df[assay] = np.nan

        for col in ["Overall Result", "Positive Assays"]:
            if col not in wide_df.columns:
                wide_df[col] = ""

        final_cols = ["Sample description 1", "Overall Result", "Positive Assays"] + desired_order
        wide_df = wide_df[final_cols]

        os.makedirs(os.path.dirname(filename) or ".", exist_ok=True)
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            wide_df.to_excel(writer, index=False, sheet_name="Results")

        wb = load_workbook(filename)
        ws = wb["Results"]

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            for col in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    continue
                val = str(cell.value).strip().lower()
                if val.startswith("positive"):
                    cell.fill = green_fill
                elif val.startswith("negative"):
                    cell.fill = red_fill
                elif val == "inconclusive":
                    cell.fill = yellow_fill

        wb.save(filename)
        return filename

    def save_full_information(self, df, filename):
        """
        Save the full-information long-format export: one row per
        (sample, assay) with the key detailed numbers, plus the Overall
        Result merged in. Control samples (PC/NC/NTC) excluded.
        """
        extra_cols = [
            "Source_File",
            "Sample description 1",
            "Assay",
            "Target",
            "Accepted Events",
            "Positives",
            "Positives (NTC-normalized)",
            "IC Target",
            "IC Positives (NTC-normalized)",
            "Fractional Abundance",
            "Result",
        ]
        missing = [c for c in extra_cols if c not in df.columns]
        if missing:
            raise ValueError(f"Missing columns in DataFrame: {missing}")

        full_df = df[extra_cols].copy()
        full_df = full_df[
            ~full_df["Sample description 1"].astype(str).str.strip().str.upper().isin(CONTROL_SAMPLE_NAMES)
        ]

        summary_df = self.add_overall_result(
            full_df[["Sample description 1", "Assay", "Result"]]
        )
        full_df = full_df.merge(summary_df, on="Sample description 1", how="left")
        full_df = full_df.drop_duplicates()

        os.makedirs(os.path.dirname(filename) or ".", exist_ok=True)
        full_df.to_excel(filename, index=False)
        return filename
